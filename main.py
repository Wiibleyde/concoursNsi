"""TODO 
- fix the file adding function
- get the right datas
- generate charts
"""

from flask import Flask, render_template, request, g, redirect
import sqlite3
import csv
import requests
import os
import pandas as pd
import openpyxl

app = Flask(__name__)

app.config["SECRET_KEY"] = "MotDePasse"

class File:
    def __init__(self, fileName):
        """Init of File class

        Args:
            fileName (str): Name of the file
        """
        self.fileName = fileName
        if 'http' in fileName:
            self.fileName=self.download(fileName)
            if not self.fileName:
                print("This file is not a csv file")

    def isCsv(self,fileName=None):
        """Test if a file is a csv file

        Args:
            fileName (str, optional): name of the file. Defaults to None.

        Returns:
            bool: true if the file is a csv
        """
        if fileName is None:
            fileName=self.fileName
        try:
            print(fileName)
            with open(fileName, 'r') as f:
                first_line = f.readline()
                if ',' in first_line or ';' in first_line:
                    return True
                else:
                    return False
        except UnicodeDecodeError or FileNotFoundError:
            return False

    def isXlsx(self, fileName=None):
        """Test if a file is a xlsx file

        Args:
            fileName (str, optional): name of the file. Defaults to None.

        Returns:
            bool: true if the file is a xlsx
        """
        if fileName is None:
            fileName=self.fileName
        try:
            excel=openpyxl.load_workbook(fileName)
            return True
        except openpyxl.utils.exceptions.InvalidFileException or FileNotFoundError:
            return False

    def download(self,url): 
        """Download file
        
        Args:
            url (str): url of a file
        """  
        try:
            print(f'Downloading {url}')
            name = self.getFileName()
            r = requests.get(url, allow_redirects=True)
            try:
                open(f'filesTemp\\{name}', 'wb').write(r.content)
                if self.isCsv(f'filesTemp\\{name}'):
                    newName=f'files\\{name}.csv'
                    os.rename(f'filesTemp\\{name}', newName)
                    print(f'{newName} is csv saving it')
                    return newName
                elif self.isXlsx(f'filesTemp\\{name}'):
                    newName=f'files\\{name}.xlsx'
                    os.rename(f'filesTemp\\{name}', newName)
                    print(f'{newName} is xlsx saving it')
                    return newName
                else:
                    os.remove(f'filesTemp\\{name}')
                    print(f'{name} is not csv or xlsx deleting it')
                return False
            except FileNotFoundError:
                return Error_Page("File not found")
        except requests.exceptions.MissingSchema:
            print(f'{url} is not a valid url')

        return f'files\\{self.getFileName()}.xlsx'

    def getFileName(self):
        """get the name of a file

        Returns:
            str: name of the file
        """
        if 'http' in self.fileName:
            print('http found')
            return self.fileName.split('/')[-1].split('.')[0]
        else:
            return self.fileName.split('\\')[-1].split('.')[0]

    def fieldNames(self):
        """name of field in csv file

        Returns:
            lst: list of fields of the csv file
        """
        if self.isCsv():
            with open(self.fileName, 'r') as f:
                reader = csv.reader(f, delimiter=';')
                firstLine=next(reader)
                cols=firstLine #[0].split(';')
                for compteur in range(len(cols)):
                    cols[compteur]="'"+cols[compteur].replace(' ','_')+"'"
                return cols               
        elif self.isXlsx():
            excel = openpyxl.load_workbook(self.fileName)
            sheet = excel.active
            cols=[]
            firstRow=next(sheet.rows)
            for c in firstRow:
                cols.append(c.value)
            print(cols)
            for compteur in range(len(cols)):
                if cols[compteur] is not None:
                    cols[compteur]="'"+cols[compteur].replace(' ','_')+"'"
                else:
                    cols[compteur]="'"+str(compteur)+"'"
            return cols   
        else:
            print("This file is not a csv file")

    def copyToSQLite(self,dbName):
        """copying the content of csv file in a sqlite database

        Args:
            dbName (str): name of database (default name: Database.db)
        """
        if self.isCsv():
            with open(self.fileName, 'r') as f:
                reader = csv.reader(f, delimiter=';')
                con = sqlite3.connect(dbName)
                cur = con.cursor()
                con.set_trace_callback(print)
                cur.execute("DROP TABLE IF EXISTS '{}'".format(self.getFileName()))
                cur.execute("CREATE TABLE '{}' ({})".format(self.getFileName(),','.join(self.fieldNames())))
                next(reader)
                for row in reader:
                    sql="INSERT INTO '{}' VALUES (?{})".format(self.getFileName(),(len(self.fieldNames())-1)*",?")
                    cur.execute(sql,row)
                con.commit()
        elif self.isXlsx():
            excel = openpyxl.load_workbook(self.fileName)
            sheet = excel.active
            con = sqlite3.connect(dbName)
            cur = con.cursor()
            con.set_trace_callback(print)
            cur.execute("DROP TABLE IF EXISTS '{}'".format(self.getFileName()))
            cur.execute("CREATE TABLE '{}' ({})".format(self.getFileName(),','.join(self.fieldNames())))
            for r in sheet.rows:
                sql="INSERT INTO '{}' VALUES (?{})".format(self.getFileName(),(len(self.fieldNames())-1)*",?")
                cur.execute(sql,[cell.value for cell in r])
            con.commit()
        else:
            print("This file is not a csv file")
        self.deleteFile()

    def deleteFile(self):
        """deleting specific file
        """
        os.remove(self.fileName)
        print(f'{self.fileName} deleted')

    def getTitle(self,dbName):
        """get the title of the sqlite db

        Args:
            dbName (str): name of database

        Returns:
            lst: list of titles of sqlite database
        """
        con = sqlite3.connect(dbName)
        cur = con.cursor()
        # con.set_trace_callback(print)
        req="SELECT name FROM pragma_table_info('{}') ORDER BY cid".format(self.getFileName())
        lstTitle=cur.execute(req).fetchall()
        con.commit()
        titles=[]
        for ele in lstTitle:
            titles.append(ele[0])
        return titles
    
    def getData(self,column):
        """get the data of one column of the sqlite db"""
        con = sqlite3.connect("files\\Database.db")
        cur = con.cursor()
        # con.set_trace_callback(print)
        req="SELECT {} FROM '{}'".format(column,self.getFileName())
        data=cur.execute(req).fetchall()
        con.commit()
        lstData=[]
        for ele in data:
            lstData.append(ele[0])
        return lstData
    
    # def getOccurence(self,column):
    #     """get occurence of all data in a list
    #     return a tuple 
    #     format : {label:'title', value:occurence}"""
    #     lstData=self.getData(column)
    #     lstOccurence=[]
    #     for ele in lstData:
    #         if ele not in lstOccurence:
    #             lstOccurence.append(ele)
    #     lstOccurence.sort()
    #     lstOccurence.reverse()
    #     lstOccurence=[{'label':ele,'value':lstData.count(ele)} for ele in lstOccurence]
    #     return lstOccurence

    def getPieDataSum(self,columns,filter=None):
        """return a dict with column and value of pie chart

        Args:
            columns (lst): list of columns| each element can be a string or a list of string
            filter (dict, optional): dict of filter to apply to DBase. Defaults to None.

        Returns:
            dict: dict of column and value of pie chart
        """
        con = sqlite3.connect("files\\Database.db")
        con.row_factory = sqlite3.Row
        cur = con.cursor()
        # con.set_trace_callback(print)
        sum=''
        rcolumns=[]
        for column in columns:
            if isinstance(column,str):
                sum+='SUM("{}") AS "{}",'.format(column,column)
                rcolumns.append(column)
            else:
                s=''
                for c in column:
                    s+='SUM("{}")+'.format(c)
                u=column[0].rindex('_')
                sum+='{} AS "{}",'.format(s[:-1],column[0][:u])
                rcolumns.append(column[0][:u])
        req='SELECT {} FROM "{}"'.format(sum[:-1],self.getFileName())
        if filter is not None:
            where=''
            for k in filter.keys():
                where+='"{}"="{}" AND '.format(k,filter[k])
            req+=' WHERE {}'.format(where[:-4])
        print(req)
        data=cur.execute(req).fetchone()
        # con.commit()
        lstData={}
        for column in rcolumns:
            lstData[column]=int(data[column])
        return lstData

    def getColumnDistinct(self,column):
        """get the distinct value of a column"""
        con=sqlite3.connect("files\\Database.db")
        cur = con.cursor()
        # con.set_trace_callback(print)
        req='SELECT DISTINCT "{}" FROM "{}"'.format(column,self.getFileName())
        data=cur.execute(req).fetchall()
        lstData=[]
        for ele in data:
            lstData.append(ele[0])
        return lstData
    
    def isColumnNumeric(self,column):
        """check if a column contains only numeric value"""
        con=sqlite3.connect("files\\Database.db")
        cur = con.cursor()
        # con.set_trace_callback(print)
        lstData=self.getColumnDistinct(column)
        for ele in lstData:
            if ele!='' and not ele.isdecimal():
                return False
        return True
        
        

def deleteDatabase():
    """delete the database named 'Database.db'
    """
    os.remove('files\\Database.db')
    print('Database deleted')

@app.route("/", methods=["GET", "POST"])
def Data_Analyser():
    """main page"""
    return render_template("Data_Analyser.html")

@app.route("/Import_CSV", methods=["GET", "POST"])
def Import_CSV():
    """import a csv file"""
    return render_template("Import_CSV.html")

@app.route("/Selection", methods=["GET", "POST"])
def Selection():
    """select a title of csv file"""
    global fichier
    Filters = []
    FiltersValues = []
    Columns = []
    if len(request.form)==0:
        if fichier is None:
            print('redirect')
            return redirect('/Import_CSV')
        else:
            Titles = fichier.getTitle("files\\Database.db")
    else:
        if request.form.get("link") != '':
            lien = request.form.get("link")
            try:
                fichier=File(lien)
            except FileNotFoundError:
                return Error_Page("File not found")
            fichier.copyToSQLite("files\\Database.db")
            Titles = fichier.getTitle("files\\Database.db")
        elif len(request.files)!=0:
            f = request.files['select-file']
            f.save(f'files\\{f.filename}')
            fichier=File(f'files\\{f.filename}')
            fichier.copyToSQLite(f'files\\Database.db')
            Titles = fichier.getTitle("files\\Database.db")
        else:
            try:
                Titles = fichier.getTitle("files\\Database.db")
            except NameError:
                print('redirect')
                return redirect('/Import_CSV')
    for i in range(9):
        Filters.append(Titles[i])
        FiltersValues.append(fichier.getColumnDistinct(Titles[i]))
    for i in range(9, len(Titles)):
        Columns.append(Titles[i])
    # filter1 = fichier.getColumnDistinct(Titles[0])
    # filter2 = fichier.getColumnDistinct(Titles[1])
    # filter3 = fichier.getColumnDistinct(Titles[2])
    # filter4 = fichier.getColumnDistinct(Titles[3])
    # filter5 = fichier.getColumnDistinct(Titles[4])
    # filter6 = fichier.getColumnDistinct(Titles[5])
    # filter7 = fichier.getColumnDistinct(Titles[6])
    # filter8 = fichier.getColumnDistinct(Titles[7])
    # filter9 = fichier.getColumnDistinct(Titles[8])
    return render_template("Selection.html", Filters = Filters, Columns = Columns, FiltersValues = FiltersValues, filtersLen=len(Filters))

@app.route("/Show_Graph", methods=["GET", "POST"])
def Show_Graph():
    """show a graph"""
    Table = fichier.getFileName()
    Filter = request.form.get("filter")
    Column = request.form.get("column")
    # data = fichier.getData(Name)
    # print(fichier.getOccurence(Name))
    # print(fichier.getPieDataSum([['0105_humanites_litterature_et_philosophie_filles','0105_humanites_litterature_et_philosophie_garcons'],'0300_langues_litterature_et_cultures_etrangeres_et_regionales_filles','0300_langues_litterature_et_cultures_etrangeres_et_regionales_garcons'],{'rentree_scolaire':'2021','numero_etablissement':'0331503E'}))
    # print(fichier.getColumnDistinct(Name))
    # print(fichier.isColumnNumeric(Name))
    return render_template("Show_Graph.html", data=data) 

@app.route("/Error_Page", methods=["GET", "POST"])
def Error_Page(error):
    """error page"""
    return render_template("Error_Page.html", error=error)

if __name__ == "__main__":
    # fichier=File('')
    # fichier.copyToSQLite("files\\Database.db")
    # fichier.getData("Nom")    
    app.run()
    try:
        deleteDatabase()
    except FileNotFoundError:
        print("Database not found")